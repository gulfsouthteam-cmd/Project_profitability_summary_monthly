"""
QBO Project Profitability Summary Monthly — parser + Flask endpoint.

Make.com POSTs the .xlsx as multipart/form-data (field name: 'file')
to /process, this returns JSON, Make writes rows to Google Sheets.

Period parsing: scans all header rows for any cell matching "Month YYYY"
and derives period_start (1st) and period_end (last day of that month).
"""

from __future__ import annotations

import calendar
import logging
import os
import re
from datetime import date, datetime
from io import BytesIO
from typing import Optional

from flask import Flask, jsonify, request
from openpyxl import load_workbook

log = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

app = Flask(__name__)

API_KEY = os.environ.get("PIPELINE_API_KEY")

HEADER_FIRST_CELL = "Project"

_PERIOD_MONTH_ONLY = re.compile(r"^([A-Za-z]+)\s+(\d{4})$")


@app.route("/process", methods=["POST"])
def process():
    if API_KEY and request.headers.get("X-Api-Key") != API_KEY:
        return jsonify({"ok": False, "error": "unauthorized"}), 401

    upload = request.files.get("file")
    if upload is None:
        return jsonify({"ok": False, "error": "missing file"}), 400

    file_bytes = upload.read()
    if not file_bytes:
        return jsonify({"ok": False, "error": "empty file"}), 400

    try:
        parsed = parse(file_bytes)
    except ValueError as e:
        log.warning("parse failed: %s", e)
        return jsonify({"ok": False, "error": str(e)}), 422
    except Exception as e:
        log.exception("unexpected parse error")
        return jsonify({"ok": False, "error": f"parse failed: {e}"}), 500

    log.info("parsed period=%s rows=%d",
             parsed["report"]["period"], parsed["report"]["row_count"])
    return jsonify(parsed["rows"]), 200


@app.route("/", methods=["GET"])
def health():
    return jsonify({"ok": True, "service": "project-profitability-monthly-parser"}), 200


def parse(file_bytes: bytes) -> dict:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header_idx = _find_header_row(rows)
    if header_idx is None:
        raise ValueError(
            f"Could not find header row starting with '{HEADER_FIRST_CELL}'. "
            "Is this actually a Project Profitability Summary export?"
        )

    # Scan every row above the header for company, title, and period
    title = None
    company = None
    period_raw = None
    period_start = None
    period_end = None

    for i in range(header_idx):
        val = _cell(rows, i)
        if not val:
            continue
        if company is None:
            company = val
        elif title is None:
            title = val
        else:
            # Check every non-blank pre-header cell for a month pattern
            result = _parse_month_only(val)
            if result[0] is not None:
                period_raw = val
                period_start, period_end = result
                break

    log.info("company=%r title=%r period_raw=%r period_end=%r",
             company, title, period_raw, period_end)

    data_rows = []
    for raw in rows[header_idx + 1:]:
        if _is_blank(raw) or _looks_like_footer(raw):
            break
        parsed_row = _parse_data_row(raw, period_end)
        if parsed_row is not None:
            data_rows.append(parsed_row)

    return {
        "report": {
            "title": title,
            "company": company,
            "period": period_raw,
            "period_start": period_start,
            "period_end": period_end,
            "generated_at": _find_generated_at(rows),
            "row_count": len(data_rows),
        },
        "rows": data_rows,
    }


def _parse_month_only(val: str):
    """Return (period_start, period_end) for 'Month YYYY', else (None, None)."""
    m = _PERIOD_MONTH_ONLY.match(val.strip())
    if not m:
        return (None, None)
    month_str, year_str = m.groups()
    try:
        dt = datetime.strptime(f"{month_str} {year_str}", "%B %Y")
        last_day = calendar.monthrange(dt.year, dt.month)[1]
        start = date(dt.year, dt.month, 1)
        end = date(dt.year, dt.month, last_day)
        return (start.isoformat(), end.isoformat())
    except ValueError:
        return (None, None)


def _cell(rows: list, idx: int) -> Optional[str]:
    if idx >= len(rows) or not rows[idx]:
        return None
    val = rows[idx][0]
    return str(val).strip() if val is not None else None


def _find_header_row(rows: list) -> Optional[int]:
    for i, row in enumerate(rows):
        if row and row[0] == HEADER_FIRST_CELL:
            return i
    return None


def _is_blank(row: tuple) -> bool:
    return not row or all(c is None or c == "" for c in row)


def _looks_like_footer(row: tuple) -> bool:
    if not row or row[0] is None:
        return False
    first = str(row[0]).strip()
    others_empty = all(c is None or c == "" for c in row[1:])
    return others_empty and bool(re.search(r"\d{4}.*(AM|PM)", first))


def _parse_data_row(row: tuple, period_end: Optional[str] = None) -> Optional[dict]:
    if len(row) < 6 or row[0] is None:
        return None
    project = str(row[0]).strip()
    customer = str(row[1]).strip() if row[1] is not None else None
    return {
        "Job Number": _extract_job_number(project),
        "Project": project,
        "Customer": customer,
        "Income__TOTAL": _to_float(row[2]),
        "Costs__TOTAL": _to_float(row[3]),
        "Profit__TOTAL": _to_float(row[4]),
        "Profit margin": _to_float(row[5]),
        "period_end": period_end,
    }


def _to_float(val) -> Optional[float]:
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


_JOB_NUMBER_TRAILING = re.compile(r"[-#]\s*(\d{3,5})\s*$")
_JOB_NUMBER_LEADING  = re.compile(r"^#\s*(\d{3,5})\b")


def _extract_job_number(project: str) -> Optional[str]:
    m = _JOB_NUMBER_TRAILING.search(project)
    if m:
        return m.group(1)
    m = _JOB_NUMBER_LEADING.search(project)
    if m:
        return m.group(1)
    return None


def _find_generated_at(rows: list) -> Optional[str]:
    for row in reversed(rows):
        if _looks_like_footer(row):
            raw = str(row[0]).strip()
            stripped = re.sub(r"^[A-Za-z]+,\s*", "", raw)
            tz_match = re.search(r"GMT([+-]\d{2}:?\d{2})\s*$", stripped)
            tz = tz_match.group(1) if tz_match else None
            stripped = re.sub(r"\s*GMT[+-]\d{2}:?\d{2}\s*$", "", stripped)
            try:
                dt = datetime.strptime(stripped, "%B %d, %Y %I:%M %p")
                iso = dt.isoformat()
                if tz:
                    if ":" not in tz:
                        tz = tz[:3] + ":" + tz[3:]
                    iso += tz
                return iso
            except ValueError:
                return raw
    return None


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port)
